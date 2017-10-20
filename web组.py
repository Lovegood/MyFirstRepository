import xml.etree.ElementTree as ET
from openpyxl import Workbook

def category(str):
    result = 'unknow'
    if str.find('code execution') != -1:
        result = '代码执行'
    elif str.find('XSS') != -1:
        result = 'XSS'
    elif str.find('CSRF') != -1:
        result = 'CSRF'
    elif str.find('SQL injection') != -1:
        result = 'SQL注入'
    elif str.find('denial of service') != -1:
        result = 'DOS'
    return result

def harm(str):
    result = ''
    if str.find('gain privileges') != -1:
        result = '可提权 '
    if str.find('cause a denial of service') != -1:
        result += '造成DOS '
    if str.find('information disclosure') != -1:
        result += '造成信息泄露 '
    return result

def main():
    wb = Workbook() #create a workbook
    ws = wb.get_active_sheet() #调用正在运行的工作表 worksheet
    ws.title = 'raw'
    rowCnt = 2

    xmlPath = 'allitems-cvrf-year-2017.xml'
    xmlTagPrefix = '{http://www.icasi.org/CVRF/schema/vuln/1.1}'
    tree = ET.ElementTree(file=xmlPath)
    root = tree.getroot()
    for child in root:
        if child.tag == xmlTagPrefix+'Vulnerability':
            dataNode = child
            CVEtitle = dataNode.find(xmlTagPrefix+'Title').text
            description = dataNode.find(xmlTagPrefix+'Notes').find(xmlTagPrefix+'Note').text
            if description.find('** RESERVED **') != -1:
                continue;
            ws.cell(row = rowCnt, column = 1).value = CVEtitle
            ws.cell(row = rowCnt, column = 2).value = category(description)
            ws.cell(row = rowCnt, column = 3).value = harm(description)
            ws.cell(row = rowCnt, column = 4).value = description
            rowCnt += 1

    wb.save('raw.xlsx') # save

if __name__ == '__main__':
    main()
